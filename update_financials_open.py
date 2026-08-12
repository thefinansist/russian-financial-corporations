#!/usr/bin/env python3
"""Обновление Excel без токена: MOEX ISS + публичные таблицы Smart-Lab."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import time
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook


DEFAULT_CONFIG = Path(__file__).with_name("companies.json")


def default_input_path() -> Path:
    """Ищет книгу в текущей папке и в локальной структуре проекта."""
    filename = "Финансовые организации.xlsx"
    candidates = [
        Path.cwd() / filename,
        Path(__file__).resolve().parent / filename,
        Path(__file__).resolve().parents[2] / filename,
    ]
    return next((path for path in candidates if path.exists()), candidates[0])
MOEX_URL = (
    "https://iss.moex.com/iss/engines/stock/markets/shares/"
    "boards/{board}/securities/{ticker}.json"
)
TRADINGVIEW_SCAN_URL = "https://scanner.tradingview.com/russia/scan"
SMARTLAB_URLS = (
    "https://smart-lab.ru/q/{ticker}/f/q/MSFO/",
    "https://smart-lab.ru/q/{ticker}/f/y/MSFO/",
    "https://smart-lab.ru/q/{ticker}/f/q/RSBU/",
    "https://smart-lab.ru/q/{ticker}/f/y/RSBU/",
)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 Chrome/124 Safari/537.36",
    "Accept": "application/json,text/html,application/xhtml+xml",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.6",
}


class SourceError(RuntimeError):
    pass


def clean_number(value: str) -> Optional[float]:
    text = str(value or "").replace("\xa0", " ").strip()
    if not text or text in {"—", "-", "?", "n/a", "N/A"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("%", "").replace("×", "").replace("x", "")
    text = text.replace(" ", "").replace(",", ".")
    text = re.sub(r"[^0-9.+-]", "", text)
    try:
        result = float(text)
    except ValueError:
        return None
    return -result if negative else result


def table_rows(payload: Dict[str, Any], block: str) -> List[Dict[str, Any]]:
    section = payload.get(block) or {}
    columns = section.get("columns") or []
    return [dict(zip(columns, row)) for row in section.get("data") or []]


def get_json(session: requests.Session, url: str, params: Dict[str, str]) -> Dict[str, Any]:
    last_error = None
    for attempt in range(2):
        try:
            response = session.get(url, params=params, timeout=(7, 18))
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < 1:
                time.sleep(2 ** attempt)
    raise SourceError("MOEX ISS не ответил: {}".format(last_error))


def fetch_moex(
    session: requests.Session, ticker: str, board: str
) -> Dict[str, Optional[float]]:
    payload = get_json(
        session,
        MOEX_URL.format(board=board, ticker=ticker),
        {
            "iss.meta": "off",
            "iss.only": "securities,marketdata",
            "securities.columns": "SECID,ISSUESIZE",
            "marketdata.columns": "SECID,LAST,MARKETPRICE,MARKETPRICETODAY,"
            "LCURRENTPRICE,LEGALCLOSEPRICE,CAPITALIZATION",
        },
    )
    securities = table_rows(payload, "securities")
    marketdata = table_rows(payload, "marketdata")
    security = securities[0] if securities else {}
    market = marketdata[0] if marketdata else {}
    price = next(
        (
            float(market[key])
            for key in (
                "LAST",
                "MARKETPRICETODAY",
                "MARKETPRICE",
                "LCURRENTPRICE",
                "LEGALCLOSEPRICE",
            )
            if market.get(key) is not None
        ),
        None,
    )
    return {
        "price": price,
        "issue_size": float(security["ISSUESIZE"])
        if security.get("ISSUESIZE") is not None
        else None,
        "market_cap_bln": float(market["CAPITALIZATION"]) / 1_000_000_000.0
        if market.get("CAPITALIZATION") is not None
        else None,
    }


def fetch_tradingview(
    session: requests.Session, tickers: List[str]
) -> Dict[str, Dict[str, Optional[float]]]:
    """Пакетная резервная котировка TradingView (неофициальный endpoint)."""
    payload = {
        "symbols": {
            "tickers": ["RUS:" + ticker for ticker in tickers],
            "query": {"types": []},
        },
        "columns": ["name", "close", "currency", "market_cap_basic"],
    }
    try:
        response = session.post(
            TRADINGVIEW_SCAN_URL,
            json=payload,
            timeout=(7, 20),
            headers={
                "Origin": "https://www.tradingview.com",
                "Referer": "https://www.tradingview.com/",
            },
        )
        response.raise_for_status()
        raw = response.json()
    except (requests.RequestException, ValueError) as exc:
        raise SourceError("TradingView не ответил: {}".format(exc))
    result: Dict[str, Dict[str, Optional[float]]] = {}
    for item in raw.get("data", []):
        symbol = str(item.get("s", ""))
        values = item.get("d") or []
        if len(values) < 4 or not symbol.startswith("RUS:"):
            continue
        ticker = symbol.split(":", 1)[1]
        price = float(values[1]) if values[1] is not None else None
        currency = str(values[2] or "")
        if price is None or price <= 0 or currency != "RUB":
            continue
        result[ticker] = {
            "price": price,
            "issue_size": None,
            "market_cap_bln": float(values[3]) / 1_000_000_000.0
            if values[3] is not None
            else None,
        }
    return result


class SmartLabTableParser(HTMLParser):
    def __init__(self) -> None:
        HTMLParser.__init__(self, convert_charrefs=True)
        self.result: Dict[str, Optional[float]] = {}
        self.field: Optional[str] = None
        self.values: List[Optional[float]] = []
        self.in_td = False
        self.skip_td = False
        self.td_text: List[str] = []

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        attributes = dict(attrs)
        if tag == "tr" and attributes.get("field"):
            self.field = str(attributes["field"]).strip()
            self.values = []
        elif tag == "td" and self.field is not None:
            self.in_td = True
            self.td_text = []
            self.skip_td = "ltm_spc" in str(attributes.get("class", "")).split()

    def handle_data(self, data: str) -> None:
        if self.in_td:
            self.td_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "td" and self.in_td:
            if not self.skip_td:
                parsed = clean_number(" ".join(self.td_text))
                self.values.append(parsed)
            self.in_td = False
            self.skip_td = False
            self.td_text = []
        elif tag == "tr" and self.field is not None:
            self.result[self.field] = self.values[-1] if self.values else None
            self.field = None
            self.values = []


def extract_smartlab_fields(html: str) -> Dict[str, Optional[float]]:
    parser = SmartLabTableParser()
    parser.feed(html)
    return parser.result


def fetch_smartlab(
    session: requests.Session, ticker: str
) -> Tuple[Dict[str, Optional[float]], str]:
    required = ("net_income", "capital")
    best: Dict[str, Optional[float]] = {}
    best_url = ""
    for template in SMARTLAB_URLS:
        url = template.format(ticker=ticker)
        try:
            response = session.get(url, timeout=(10, 25))
            response.raise_for_status()
        except requests.RequestException:
            continue
        fields = extract_smartlab_fields(response.text)
        if sum(fields.get(key) is not None for key in required) > sum(
            best.get(key) is not None for key in required
        ):
            best, best_url = fields, url
        if all(fields.get(key) is not None for key in required):
            return fields, url
    if not best:
        raise SourceError("Smart-Lab не вернул таблицу для {}".format(ticker))
    return best, best_url


def load_config(path: Path) -> Dict[str, Dict[str, str]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {str(name).strip().casefold(): spec for name, spec in raw.items()}


def build_company_data(
    session: requests.Session,
    name: str,
    spec: Dict[str, str],
    moex_enabled: bool = True,
    prefetched_market: Optional[Dict[str, Optional[float]]] = None,
    market_source: str = "MOEX ISS",
) -> Tuple[Dict[str, Optional[float]], List[str], str]:
    ticker = spec["ticker"].upper()
    board = spec.get("class_code", "TQBR").upper()
    warnings: List[str] = []
    market: Dict[str, Optional[float]] = {}
    smart: Dict[str, Optional[float]] = {}
    smart_url = ""

    if prefetched_market is not None:
        market = prefetched_market
    elif moex_enabled:
        try:
            market = fetch_moex(session, ticker, board)
        except SourceError as exc:
            warnings.append("{}: {}".format(name, exc))
    try:
        smart, smart_url = fetch_smartlab(session, spec.get("smartlab_ticker", ticker))
    except SourceError as exc:
        warnings.append("{}: {}".format(name, exc))

    common_shares_mln = smart.get("number_of_shares")
    pref_shares_mln = smart.get("number_of_priv_shares") or 0.0
    # Не смешиваем старую цену из книги с новым числом акций после сплита/допэмиссии.
    # Если MOEX не дал актуальную цену, оба рыночных входа сохраняются из исходника.
    shares_mln = None
    if market.get("price") is not None:
        shares_mln = (
            common_shares_mln + pref_shares_mln
            if common_shares_mln is not None
            else (
                market.get("issue_size") / 1_000_000.0
                if market.get("issue_size") is not None
                else None
            )
        )
    net_income = smart.get("net_income")
    # У нефинансовых компаний собственный капитал размечен как book_value.
    equity = smart.get("capital")
    if equity is None:
        equity = smart.get("book_value")
    if equity is None:
        equity = smart.get("net_assets")
    payout = smart.get("div_payout_ratio")
    roe = smart.get("roe")
    if equity is not None and equity <= 0:
        roe = None
    # При убытке, почти нулевой прибыли или экстремальном отношении payout неинформативен.
    if net_income is None or net_income <= 0 or payout is None or payout < 0 or payout > 500:
        payout = None
    return (
        {
            "price": market.get("price"),
            "shares_mln": shares_mln,
            "net_income_bln": net_income,
            "equity_bln": equity,
            "roe": roe / 100.0 if roe is not None else None,
            "payout": payout / 100.0 if payout is not None else None,
            "reference_market_cap_bln": market.get("market_cap_bln"),
        },
        warnings,
        "{}; {}".format(market_source, smart_url)
        if market.get("price") is not None
        else smart_url,
    )


def update_workbook(
    source: Path,
    destination: Path,
    sheet_name: str,
    config: Dict[str, Dict[str, str]],
    clear_missing: bool,
) -> Tuple[List[Dict[str, str]], List[str]]:
    workbook = load_workbook(source)
    if sheet_name not in workbook.sheetnames:
        raise ValueError("В книге нет листа {!r}".format(sheet_name))
    sheet = workbook[sheet_name]
    session = requests.Session()
    session.headers.update(HEADERS)
    audit: List[Dict[str, str]] = []
    warnings: List[str] = []
    sheet["A1"] = date.today()
    sheet["A1"].number_format = "dd.mm.yy"

    companies = []
    for row_number in range(2, sheet.max_row + 1):
        company = str(sheet.cell(row_number, 1).value or "").strip()
        if company:
            companies.append((row_number, company))

    moex_enabled = True
    prefetched_by_ticker: Dict[str, Dict[str, Optional[float]]] = {}
    market_source = "MOEX ISS"
    first_spec = next((config.get(company.casefold()) for _, company in companies if config.get(company.casefold())), None)
    if first_spec:
        print("Проверяю доступность MOEX ISS...", flush=True)
        try:
            first_ticker = first_spec["ticker"].upper()
            prefetched_by_ticker[first_ticker] = fetch_moex(
                session, first_ticker, first_spec.get("class_code", "TQBR").upper()
            )
        except SourceError as exc:
            moex_enabled = False
            warnings.append(str(exc) + "; используется резервный источник TradingView")
            print("MOEX ISS недоступен — запрашиваю резервные цены TradingView...", flush=True)
            market_source = "TradingView (резервный неофициальный endpoint)"
            tickers = [
                config[company.casefold()]["ticker"].upper()
                for _, company in companies
                if company.casefold() in config
            ]
            try:
                prefetched_by_ticker = fetch_tradingview(session, tickers)
                print(
                    "TradingView: получено цен {}/{}".format(len(prefetched_by_ticker), len(tickers)),
                    flush=True,
                )
            except SourceError as tv_exc:
                warnings.append(str(tv_exc) + "; старые цены сохранены")
                market_source = "нет свежей котировки"
                print("TradingView также недоступен — старые цены будут сохранены.", flush=True)

    for position, (row_number, company) in enumerate(companies, 1):
        spec = config.get(company.casefold())
        if spec is None:
            print("[{}/{}] {}: нет в companies.json".format(position, len(companies), company), flush=True)
            audit.append({"company": company, "status": "пропущено", "details": "нет в companies.json", "source": ""})
            continue
        print("[{}/{}] Обновляю {} ({})...".format(position, len(companies), company, spec["ticker"]), flush=True)
        ticker = spec["ticker"].upper()
        values, company_warnings, source_url = build_company_data(
            session,
            company,
            spec,
            moex_enabled=moex_enabled,
            prefetched_market=prefetched_by_ticker.get(ticker),
            market_source=market_source,
        )
        warnings.extend(company_warnings)
        reference_cap = values["reference_market_cap_bln"]
        implied_cap = (
            values["price"] * values["shares_mln"] / 1000.0
            if values["price"] is not None and values["shares_mln"] is not None
            else None
        )
        if reference_cap and implied_cap and abs(implied_cap / reference_cap - 1) > 0.20:
            warnings.append(
                "{}: проверка капитализации не пройдена: расчет {:.1f} млрд ₽, источник {:.1f} млрд ₽"
                .format(company, implied_cap, reference_cap)
            )
        targets = {
            2: values["price"],
            3: values["shares_mln"],
            4: values["net_income_bln"],
            5: values["equity_bln"],
            9: values["roe"],
            11: values["payout"],
        }
        updated = []
        missing = []
        for column, value in targets.items():
            if value is None:
                missing.append(sheet.cell(1, column).value)
                if clear_missing:
                    sheet.cell(row_number, column).value = None
            else:
                sheet.cell(row_number, column).value = value
                updated.append(str(sheet.cell(1, column).value))

        sheet.cell(row_number, 6).value = '=IFERROR(B{0}*C{0}/1000,"")'.format(row_number)
        sheet.cell(row_number, 7).value = '=IFERROR(F{0}/D{0},"")'.format(row_number)
        sheet.cell(row_number, 8).value = '=IFERROR(F{0}/E{0},"")'.format(row_number)
        if values["roe"] is None and values["equity_bln"] is not None and values["equity_bln"] > 0:
            sheet.cell(row_number, 9).value = '=IFERROR(D{0}/E{0},"")'.format(row_number)
        elif values["equity_bln"] is not None and values["equity_bln"] <= 0:
            sheet.cell(row_number, 9).value = None
        if values["equity_bln"] is not None and values["equity_bln"] <= 0:
            sheet.cell(row_number, 10).value = None
        else:
            sheet.cell(row_number, 10).value = '=IFERROR(I{0}*(1-K{0}),"")'.format(row_number)
        for column in (2, 3, 4, 5, 6):
            sheet.cell(row_number, column).number_format = '#,##0.0;[Red](#,##0.0);-'
        if values["net_income_bln"] is not None and abs(values["net_income_bln"]) < 1:
            sheet.cell(row_number, 4).number_format = '#,##0.000;[Red](#,##0.000);-'
        for column in (7, 8):
            sheet.cell(row_number, column).number_format = '0.0x;[Red](0.0x);-'
        for column in (9, 10, 11):
            sheet.cell(row_number, column).number_format = '0.0%;[Red](0.0%);-'
        details = "обновлено: {}".format(", ".join(updated))
        if missing:
            details += "; сохранены старые: {}".format(", ".join(map(str, missing)))
        audit.append({"company": company, "status": "обновлено", "details": details, "source": source_url})

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return audit, warnings


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Обновление финансовой таблицы без токена")
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=default_input_path(),
        help="Исходный файл .xlsx (по умолчанию: %(default)s)",
    )
    parser.add_argument("-o", "--output", type=Path, help="Путь для результата")
    parser.add_argument("--sheet", default="Лист1", help="Имя листа")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG, help="Файл companies.json")
    parser.add_argument("--in-place", action="store_true", help="Обновить исходный файл с резервной копией")
    parser.add_argument("--clear-missing", action="store_true", help="Очищать значения, отсутствующие в источниках")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        print("Ошибка: файл не найден: {}".format(args.input), file=sys.stderr)
        return 2
    if args.in_place:
        destination = args.input
        backup = args.input.with_suffix(args.input.suffix + ".bak")
        shutil.copy2(args.input, backup)
        print("Резервная копия: {}".format(backup), flush=True)
    else:
        destination = args.output or args.input.with_name(args.input.stem + "_обновлено_open.xlsx")
    try:
        audit, warnings = update_workbook(
            args.input,
            destination,
            args.sheet,
            load_config(args.config),
            args.clear_missing,
        )
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print("Ошибка: {}".format(exc), file=sys.stderr)
        return 1

    log_path = destination.with_suffix(".log.csv")
    with log_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["company", "status", "details", "source"], delimiter=";")
        writer.writeheader()
        writer.writerows(audit)
        for warning in warnings:
            writer.writerow({"company": "", "status": "предупреждение", "details": warning, "source": ""})
    print("Готово: {}".format(destination), flush=True)
    print("Журнал: {}".format(log_path), flush=True)
    if warnings:
        print("Предупреждений: {}".format(len(warnings)), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
